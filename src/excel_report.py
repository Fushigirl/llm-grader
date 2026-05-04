# coding: utf-8
"""
Convert a grades CSV to a formatted Excel workbook.

Usage:
  python excel_report.py <grades.csv>
  python excel_report.py <grades.csv> --assignment-file data/assignment.xlsx --rubric-file data/rubric.txt

Called automatically by grade.py unless --no-excel is passed.
"""
import re
import sys
import argparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLOR_MARU   = "C6EFCE"  # green  ○
COLOR_DELTA  = "FFEB9C"  # yellow △
COLOR_BATU   = "FFC7CE"  # red    ×
COLOR_ERROR  = "D9D9D9"  # grey   unknown
COLOR_HEADER = "4472C4"  # blue header
COLOR_ROW_ODD  = "FFFFFF"
COLOR_ROW_EVEN = "F2F7FF"

_THIN   = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

DEFAULT_ASSIGNMENT_FILE = "data/assignment.xlsx"
DEFAULT_RUBRIC_FILE = "data/rubric.txt"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _score_color(cell_value: str) -> str:
    if not isinstance(cell_value, str):
        return COLOR_ERROR
    if "○" in cell_value:
        return COLOR_MARU
    if "△" in cell_value:
        return COLOR_DELTA
    if "×" in cell_value:
        return COLOR_BATU
    return COLOR_ERROR


def _extract_symbol(text: str) -> str:
    if not isinstance(text, str):
        return "?"
    m = re.search(r"[○△×]", text)
    return m.group() if m else "?"


def _shorten_name(name: str) -> str:
    """'田中　太郎 TANAKA Taro_...' → '田中　太郎'"""
    if not isinstance(name, str):
        return name
    m = re.match(r"^([^\s_]+(?:\s+[^\s_]+)?)", name)
    return m.group(1).strip() if m else name


def _clean_comment(text: str, prefix: str) -> str:
    """'基準1: ○ ― コメント' → 'コメント'"""
    if not isinstance(text, str):
        return ""
    text = re.sub(rf"^{prefix}\s*[:：]\s*[○△×]\s*[―\-–—]\s*", "", text.strip())
    return text.strip()


def _clean_souhy(text: str) -> str:
    if not isinstance(text, str):
        return ""
    return re.sub(r"^総評\s*[:：]?\s*", "", text.strip()).strip()


def _normalize(s: str) -> str:
    return re.sub(r"[\s　]+", "", str(s).strip())


def _load_assignment(assignment_file: str) -> list:
    """Return [(normalized_name, original_name, evaluator), ...] or []."""
    try:
        df = pd.read_excel(assignment_file)
        return [
            (_normalize(r["提出者　氏名"]),
             str(r["提出者　氏名"]).strip(),
             str(r["採点者"]).strip())
            for _, r in df.iterrows()
        ]
    except Exception:
        return []


def _match_evaluator(raw_name: str, assignment: list) -> str:
    norm = _normalize(raw_name)
    for key_norm, _, evaluator in assignment:
        if key_norm and key_norm in norm:
            return evaluator
    return ""


def _assignment_order(raw_name: str, assignment: list) -> int:
    norm = _normalize(raw_name)
    for i, (key_norm, _, _) in enumerate(assignment):
        if key_norm and key_norm in norm:
            return i
    return 99999


# ---------------------------------------------------------------------------
# Core
# ---------------------------------------------------------------------------

def make_excel(
    csv_path: str,
    out_path: str,
    assignment_file: str = DEFAULT_ASSIGNMENT_FILE,
    rubric_file: str = DEFAULT_RUBRIC_FILE,
) -> None:
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    assignment = _load_assignment(assignment_file)

    df["_order"] = df["学生"].apply(lambda n: _assignment_order(str(n), assignment))
    df = df.sort_values("_order").reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "採点結果"

    # Detect criteria columns dynamically (基準1, 基準2, 基準3, ...)
    criteria_nums = sorted(
        int(re.search(r"\d+", c).group())
        for c in df.columns
        if re.match(r"^基準\d+$", c)
    )

    columns = [("No.", 5, "center"), ("学生名", 18, "left"), ("採点者", 12, "left")]
    for n in criteria_nums:
        columns.append((f"基準{n}",         6,  "center"))
        columns.append((f"基準{n} 判断理由", 40, "left"))
    columns += [("総評", 50, "left"), ("抽出した考察", 55, "left")]

    # Header row
    for c, (h, w, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 24

    # Data rows
    for i, row in df.iterrows():
        r = i + 2
        bg = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN

        criteria_raws = {n: str(row.get(f"基準{n}", "")) for n in criteria_nums}
        souhy     = _clean_souhy(str(row.get("総評", "")))
        extracted = str(row.get("抽出考察", ""))
        raw_name  = str(row.get("学生", ""))
        name      = _shorten_name(raw_name)
        evaluator = _match_evaluator(raw_name, assignment)

        row_data = [i + 1, name, evaluator]
        for n in criteria_nums:
            row_data.append(_extract_symbol(criteria_raws[n]))
            row_data.append(_clean_comment(criteria_raws[n], rf"基準\s*{n}"))
        row_data += [souhy, extracted]
        row_colors = [bg, bg, bg]
        for n in criteria_nums:
            row_colors.append(_score_color(criteria_raws[n]))
            row_colors.append(bg)
        row_colors += [bg, bg]

        row_aligns = [al for (_, _, al) in columns]

        for c, (v, fc, al) in enumerate(zip(row_data, row_colors, row_aligns), 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=fc)
            cell.alignment = Alignment(horizontal=al, vertical="top", wrap_text=True)
            cell.border = _BORDER
            cell.font = Font(size=10)

        # 基準N 記号列 — larger, centred
        symbol_cols = [4 + i * 2 for i in range(len(criteria_nums))]
        for c in symbol_cols:
            ws.cell(row=r, column=c).font = Font(bold=True, size=14)
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        ws.row_dimensions[r].height = 90

    ws.freeze_panes = "A2"

    # Sheet 2: rubric
    ws2 = wb.create_sheet("採点基準")
    try:
        rubric = open(rubric_file, encoding="utf-8").read()
    except Exception:
        rubric = f"Rubric file not found: {rubric_file}"

    ws2.column_dimensions["A"].width = 100
    cell = ws2.cell(row=1, column=1, value="採点基準")
    cell.font = Font(bold=True, color="FFFFFF", size=13)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER
    ws2.row_dimensions[1].height = 24

    cell = ws2.cell(row=2, column=1, value=rubric)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.font = Font(size=11)
    cell.border = _BORDER
    ws2.row_dimensions[2].height = 200

    wb.save(out_path)
    print(f"Saved Excel: {out_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Convert grades CSV to Excel")
    parser.add_argument("csv", help="Path to grades CSV file")
    parser.add_argument("--out", type=str, help="Output Excel path (default: same name as CSV)")
    parser.add_argument("--assignment-file", default=DEFAULT_ASSIGNMENT_FILE)
    parser.add_argument("--rubric-file", default=DEFAULT_RUBRIC_FILE)
    args = parser.parse_args()

    out_path = args.out or args.csv.replace(".csv", ".xlsx")
    make_excel(args.csv, out_path, args.assignment_file, args.rubric_file)


if __name__ == "__main__":
    main()
